import sqlite3
from sqlite3 import Error
import datetime
import csv
import sys
import re
import openpyxl as xl
f_fecha="^[0-9]{2}[/][0-9]{2}[/][0-9]{4}$"
libros={}
libros_clave=[]
obras_autor=[]
obras_genero=[]
obras_año=[]
reporte_filtro=[]

nombre=input('Escriba el nombre que llevara el archivo:\n').lower()
nombre_csv=(f"{nombre}.csv")
nombre_xl=(f"{nombre}.xlsx")
nombre_bd=(f"{nombre}.db")

try:
    with open(nombre_csv,"r", newline="") as archivo:
        lector = csv.reader(archivo)
        next(lector)
        for clave, titulo, a_autor, n_autor, genero, año, isbn, fecha in lector:
            libros[int(clave)]=(titulo, a_autor, n_autor, genero, int(año), int(isbn), fecha)
            libros_clave.append((clave, titulo, a_autor, n_autor, genero, int(año), int(isbn), fecha))
    try:
        with sqlite3.connect(nombre_bd) as conn:
            cursor = conn.cursor()
            cursor.execute("CREATE TABLE IF NOT EXISTS autor \
            (autor_id INTEGER PRIMARY KEY, apellido TEXT NOT NULL, nombre TEXT NOT NULL);")
            cursor.execute("CREATE TABLE IF NOT EXISTS genero \
            (genero_id INTEGER PRIMARY KEY, nombre TEXT NOT NULL);")
            cursor.execute("CREATE TABLE IF NOT EXISTS libros \
            (libro_id INTEGER PRIMARY KEY, titulo TEXT NOT NULL, autor_id INTEGER NOT NULL, genero_id INTEGER NOT NULL, año INTEGER NOT NULL, isbn INTEGER NOT NULL, fecha TEXT NOT NULL,\
            FOREIGN KEY(genero_id) REFERENCES genero(genero_id), FOREIGN KEY(autor_id) REFERENCES autor(autor_id));")
            print("Tablas creadas exitosamente")
    except Error as e:
        print (e)
    except:
        print(f"Se produjo el siguiente error: {sys.exc_info()[0]}")
    finally:
        conn.close()
except FileNotFoundError:
    print("El archivo no se encontró, se procede a trabajar con un conjunto vacío")

    try:
        with sqlite3.connect(nombre_bd) as conn:
            cursor = conn.cursor()
            cursor.execute("CREATE TABLE IF NOT EXISTS autor \
            (autor_id INTEGER PRIMARY KEY, apellido TEXT NOT NULL, nombre TEXT NOT NULL);")
            cursor.execute("CREATE TABLE IF NOT EXISTS genero \
            (genero_id INTEGER PRIMARY KEY, nombre TEXT NOT NULL);")
            cursor.execute("CREATE TABLE IF NOT EXISTS libros \
            (libro_id INTEGER PRIMARY KEY, titulo TEXT NOT NULL, autor_id INTEGER NOT NULL, genero_id INTEGER NOT NULL, año INTEGER NOT NULL, isbn INTEGER NOT NULL, fecha TEXT NOT NULL,\
            FOREIGN KEY(genero_id) REFERENCES genero(genero_id), FOREIGN KEY(autor_id) REFERENCES autor(autor_id));")
            print("Tablas creadas exitosamente")
    except Error as e:
        print (e)
    except:
        print(f"Se produjo el siguiente error: {sys.exc_info()[0]}")
    finally:
        conn.close()

    while True:
        try:
            try:
                with sqlite3.connect(nombre_bd) as conn:
                    cursor = conn.cursor()
                    cursor.execute("SELECT * FROM autor")
                    registros = cursor.fetchall()

                    if registros:
                        print('Los autores disponibles son')
                        print("Claves\tApellido\tNombre")
                        print("*" * 30)
                        for clave, apellido, nombre in registros:
                            print(f"{clave:^6}\t{apellido}\t{nombre}")
                    else:
                        print("No se encontraron Autores disponibles, intenta ingresar algunos antes de ingresar libros")
            except Error as e:
                print (e)
            except Exception:
                print(f"Se produjo el siguiente error: {sys.exc_info()[0]}")
            finally:
                conn.close()

            try:
                with sqlite3.connect(nombre_bd) as conn:
                    cursor = conn.cursor()
                    cursor.execute("SELECT * FROM genero")
                    registros = cursor.fetchall()

                    if registros:
                        print('Los generos disponibles son')
                        print("Claves\tNombre")
                        print("*" * 30)
                        for clave, nombre in registros:
                            print(f"{clave:^6}\t{nombre}")
                    else:
                        print("No se encontraron Generos disponibles, intenta ingresar algunos antes de ingresar libros")
            except Error as e:
                print (e)
            except Exception:
                print(f"Se produjo el siguiente error: {sys.exc_info()[0]}")
            finally:
                conn.close()
        except Error as e:
            print (e)
        except Exception:
            print(f"Se produjo el siguiente error: {sys.exc_info()[0]}")
        finally:
            conn.close()

        opcion=int(input(f'Seleccione alguna de las siguientes opciones:\n 1:Subir libro\n 2:Consultas y Reportes\n 3:Registrar Autor\n 4:Registrar Genero\n 5:Salir\n'))

        if opcion==1:
            clave=max(libros, default=0)+1
            while True:
                titulo=input('Ingrese el titulo del libro: ').upper()
                if titulo=='':
                    print('No se permite dejar vacios. Intente de nuevo ')
                    continue
                else:
                    break
            while True:
                try:
                    with sqlite3.connect(nombre_bd) as conn:
                        cursor = conn.cursor()
                        cursor.execute("SELECT * FROM autor")
                        registros = cursor.fetchall()

                        if registros:
                            print('Los autores disponibles son')
                            print("Claves\tApellido\tNombre")
                            print("*" * 30)
                            for clave, apellido, nombre in registros:
                                print(f"{clave:^6}\t{apellido}\t{nombre}")
                        else:
                            print("No se encontraron Autores disponibles, intenta ingresar algunos antes de ingresar libros")
                            nom_autor=input('Ingresa el nombre del autor a registrar:\n').upper()
                            if nom_autor=='':
                                print('No se permite dejar vacios. Intente de nuevo')
                                continue
                            ap_autor=input('Ingresa el apellido del autor:\n').upper()
                            if ap_autor=='':
                                print('No se permite dejar vacios. Intente de nuevo')
                                continue
                            else:
                                try:
                                    with sqlite3.connect(nombre_bd) as conn:
                                        cursor = conn.cursor()
                                        valores = (ap_autor, nom_autor)
                                        cursor.execute("INSERT INTO autor (apellido, nombre) VALUES(?,?)", valores)
                                        print(f"La clave asignada fue {cursor.lastrowid}")
                                except Error as e:
                                    print (e)
                                except Exception:
                                    print(f"Se produjo el siguiente error: {sys.exc_info()[0]}")
                                finally:
                                    conn.close()
                except Error as e:
                    print (e)
                except Exception:
                    print(f"Se produjo el siguiente error: {sys.exc_info()[0]}")
                finally:
                    conn.close()
                try:
                    id_autor=int(input('Ingrese la clave del autor del libro: '))
                except ValueError:
                    print('No se permite dejar vacios. Intente de nuevo ')
                    continue
                else:
                    break
            while True:
                try:
                    with sqlite3.connect(nombre_bd) as conn:
                        cursor = conn.cursor()
                        cursor.execute("SELECT * FROM genero")
                        registros = cursor.fetchall()

                        if registros:
                            print('Los generos disponibles son')
                            print("Claves\tNombre")
                            print("*" * 30)
                            for clave, nombre in registros:
                                print(f"{clave:^6}\t{nombre}")
                        else:
                            print("No se encontraron Generos disponibles, intenta ingresar algunos antes de ingresar libros")
                            while True:
                                nom_genero=input('Ingresa el nombre del genero:\n').upper()
                                if nom_genero=='':
                                    print('No se permite dejar vacios. Intente de nuevo')
                                    continue
                                else:
                                    break
                            try:
                                with sqlite3.connect(nombre_bd) as conn:
                                    cursor = conn.cursor()
                                    valores = (nom_genero,)
                                    cursor.execute("INSERT INTO genero (nombre) VALUES(?)", valores)
                                    print(f"La clave asignada fue {cursor.lastrowid}")
                            except Error as e:
                                print (e)
                            except Exception:
                                print(f"Se produjo el siguiente error: {sys.exc_info()[0]}")
                            finally:
                                conn.close()
                except Error as e:
                    print (e)
                except:
                    print(f"Se produjo el siguiente error: {sys.exc_info()[0]}")
                finally:
                    conn.close()
                try:
                    id_genero=int(input('Ingrese la clave del genero del libro: '))
                except ValueError:
                    print('No se permite dejar vacios. Intente de nuevo ')
                    continue
                else:
                    break
            while True:
                try:
                    año=int(input('Ingrese el año en que fue publicado el libro: '))
                except ValueError:
                    print('No son validas las letras en el año ')
                    continue
                else:
                    if len(str(año)) > 4:
                        print('No existe ese. Intente de nuevo ')
                        continue
                    else:
                        break
            while True:
                try:      
                    isbn=int(input('Ingrese el ISBN del libro: '))
                except ValueError:
                    print('No son validas las letras en el ISBN')
                    continue
                else:
                    if isbn=='':
                        print('No se permite dejar vacios. Intente de nuevo')
                        continue
                    else:
                        break
            while True:
                fecha_adquisicion=input('Ingrese la fecha en que fue adquirido el libro (dd/mm/aaaa): ')
                if fecha_adquisicion=='':
                    print('No se permite dejar vacios. Intente de nuevo ')
                    continue
                if (not bool(re.match(f_fecha, fecha_adquisicion))):
                    print("La fecha inicial no tiene el formato (dd/mm/aaaa)")
                    continue
                else:
                    try:
                        fecha_adq= datetime.datetime.strptime(fecha_adquisicion, "%d/%m/%Y").date()
                    except:
                        print('La fecha ingresada no existe, ingresa una nueva')
                        continue
                    fecha = (fecha_adq.strftime("%d/%m/%Y"))
                    break
            try:
                with sqlite3.connect(nombre_bd) as conn:
                    cursor = conn.cursor()
                    libro=(titulo,id_autor,id_genero,año,isbn,fecha)
                    cursor.execute("INSERT INTO libros (titulo, autor_id, genero_id, año, isbn, fecha) VALUES(?,?,?,?,?,?)", libro)
                    print("Registro agregado exitosamente")
            except Error as e:
                print (e)
            except:
                print(f"Se produjo el siguiente error: {sys.exc_info()[0]}")
            finally:
                conn.close()
            try:
                with sqlite3.connect(nombre_bd) as conn:
                    cursor = conn.cursor()
                    cursor.execute("SELECT libros.libro_id, libros.titulo, autor.apellido, autor.nombre, genero.nombre, libros.año, libros.isbn, libros.fecha \
                                    FROM libros \
                                    INNER JOIN autor ON libros.autor_id = autor.autor_id\
                                    INNER JOIN genero ON libros.genero_id = genero.genero_id\
                                    ")
                    registro_libro = cursor.fetchall()

                    if registro_libro:
                        print("\t\tClave\t\tTitulo\t\tApellido Autor\t\tNombre Autor\t\tGenero\t\tAño de Publicación\t\tISBN\t\tFecha de Adquisición")
                        for clave, titulo, a_autor, n_autor, genero, año, isbn, fecha in registro_libro:
                            print(f"{clave}\t{titulo}\t{a_autor}\t{n_autor}\t{genero}\t{año}\t{isbn}\t{fecha}")
                            libros[clave]=[titulo,a_autor,n_autor,genero,año,isbn,fecha]
                            libros_clave.append((clave, titulo, a_autor, n_autor, genero, int(año), int(isbn), fecha))
                    else:
                        print('No se encontro ningun libro')
                        continue
            except Error as e:
                print (e)
            except:
                print(f"Se produjo el siguiente error: {sys.exc_info()[0]}")
            finally:
                conn.close()

        elif opcion==2:
            while True:
                menu=int(input(f'Seleccione alguna de las siguientes opciones:\n1:Consulta de titulo\n2:Reportes\n3:Volver al menu\n'))
                if menu==1:
                    while True:
                        consulta=int(input(f'Seleccione alguna de las siguientes Consultas:\n1:Por titulo\n2:Por ISBN\n3:Salir\n'))
                        if consulta==1:
                            try:
                                with sqlite3.connect(nombre_bd) as conn:
                                    cursor = conn.cursor()
                                    cursor.execute("SELECT titulo FROM libros")
                                    titulos_registro = cursor.fetchall()

                                    if titulos_registro:
                                        print("Titulos")
                                        for titulo in titulos_registro:
                                            print(f"{titulo}")
                                    else:
                                        print('No se encontro ningun libro')
                                        break
                            except Error as e:
                                print (e)
                                continue
                            except:
                                print(f"Se produjo el siguiente error: {sys.exc_info()[0]}")
                                continue
                            finally:
                                conn.close()
                            consulta_titulo=input('Ingrese el nombre del titulo a buscar: ').upper()
                            try:
                                with sqlite3.connect(nombre_bd) as conn:
                                    cursor = conn.cursor()
                                    valores = {"titulo":consulta_titulo}
                                    cursor.execute("SELECT libros.libro_id, libros.titulo, autor.apellido, autor.nombre, genero.nombre, libros.año, libros.isbn, libros.fecha \
                                                   FROM libros \
                                                   INNER JOIN autor ON libros.autor_id = autor.autor_id\
                                                   INNER JOIN genero ON libros.genero_id = genero.genero_id\
                                                   WHERE titulo = :titulo", valores)
                                    registro_libro = cursor.fetchall()

                                    if registro_libro:
                                        print("Clave\t\tTitulo\t\tApellido Autor\t\tNombre Autor\t\tGenero\t\tAño de Publicación\t\tISBN\t\tFecha de Adquisición")
                                        for clave, titulo, a_autor, n_autor, genero, año, isbn, fecha in registro_libro:
                                            print(f"{clave}\t{titulo}\t{a_autor}\t{n_autor}\t{genero}\t{año}\t{isbn}\t{fecha}")
                                    else:
                                        print('No se encontro ningun libro con ese nombre')
                                        continue
                            except Error as e:
                                print (e)

                            except:
                                print(f"Se produjo el siguiente error: {sys.exc_info()[0]}")
                            finally:
                                conn.close()

                        elif consulta==2:
                            try:
                                with sqlite3.connect(nombre_bd) as conn:
                                    cursor = conn.cursor()
                                    cursor.execute("SELECT isbn FROM libros")
                                    isbn_registro = cursor.fetchall()

                                    if isbn_registro:
                                        print("ISBN")
                                        for isbn in isbn_registro:
                                            print(f"{isbn}")
                                    else:
                                        print('No se encontro ningun libro')
                                        break
                            except Error as e:
                                print (e)
                                continue
                            except:
                                print(f"Se produjo el siguiente error: {sys.exc_info()[0]}")
                                continue
                            finally:
                                conn.close()
                            consulta_isbn=int(input('Ingrese el isbn a buscar: '))
                            try:
                                with sqlite3.connect(nombre_bd) as conn:
                                    cursor = conn.cursor()
                                    valores = {"isbn":consulta_isbn}
                                    cursor.execute("SELECT libros.libro_id, libros.titulo, autor.apellido, autor.nombre, genero.nombre, libros.año, libros.isbn, libros.fecha \
                                                   FROM libros \
                                                   INNER JOIN autor ON libros.autor_id = autor.autor_id\
                                                   INNER JOIN genero ON libros.genero_id = genero.genero_id\
                                                   WHERE isbn = :isbn", valores)
                                    registro_libro = cursor.fetchall()

                                    if registro_libro:
                                        print("Clave\t\tTitulo\t\tApellido Autor\t\tNombre Autor\t\tGenero\t\tAño de Publicación\t\tISBN\t\tFecha de Adquisición")
                                        for clave, titulo, a_autor, n_autor, genero, año, isbn, fecha in registro_libro:
                                            print(f"{clave}\t{titulo}\t{a_autor}\t{n_autor}\t{genero}\t{año}\t{isbn}\t{fecha}")
                                    else:
                                        print('No se encontro ningun libro con ese isbn')
                                        continue
                            except Error as e:
                                print (e)
                                continue
                            except:
                                print(f"Se produjo el siguiente error: {sys.exc_info()[0]}")
                                continue
                            finally:
                                conn.close()
                        elif consulta==3:
                            break
                        else:
                            print('Opcion no valida')
                            continue

                elif menu==2:
                    while True:
                        reportes=int(input(f"Seleccione por que medio desea realizar los reportes:\n1:Ver catalogo completo\n2:Por autor\n3:Genero\n4:Año publicacion\n5:Volver al menu consultas y reportes\n"))
                        if reportes==1:
                            diccionario = libros.values()
                            print("\n** Catálogo completo ** ")
                            print("Titulo\t\t\tApellido Autor\t\t\tNombre Autor\t\t\tGenero\t\t\tAño de Publicación\t\t\tISBN\t\t\tFecha de Adquisición")
                            print("*" * 120)
                            for titulo, a_autor, n_autor, genero, año_publicacion, isbn, fecha in diccionario:
                                print(f"{titulo: <15} | {a_autor: <15} | {n_autor: <15} | {genero: <15} | {año_publicacion: <15} | {isbn:<15} |  {fecha:<15}")
                        
                            while True:
                                exportar=int(input('Si quiere exportar seleccione: \n1: CSV \n2: Excel \n3:No exportar y salir\n'))
                                if exportar==1:
                                    with open(nombre_csv,'w',newline='') as archivo:
                                        grabador = csv.writer(archivo)
                                        grabador.writerow(("Clave", "Titulo", "Apellido Autor", "Nombre Autor" "Genero", "Año", "ISBN", "Fecha de adquisicion"))
                                        grabador.writerows([(clave, datos[0], datos[1], datos[2], datos[3], datos[4], datos[5], datos[6]) for clave, datos in libros.items()])
                                    print(f"El archivo fue guardado con el nombre: {nombre_csv}")
                                elif exportar==2:
                                    nombre_xl = xl.Workbook()
                                    
                                    hoja = nombre_xl["Sheet"] 
                                    hoja.title = "Reporte Completo"
                                    
                                    hoja["A1"].value = "Clave"
                                    hoja["B1"].value = "Titulo"
                                    hoja["C1"].value = "Apellido Autor"
                                    hoja["D1"].value = "Nombre Autor"
                                    hoja["E1"].value = "Genero"
                                    hoja["F1"].value = "Año"
                                    hoja["G1"].value = "ISBN"
                                    hoja["H1"].value = "Fecha de adquisicion"
                                    
                                    renglon = 2
                                    for clave, datos in libros.items():
                                        titulo, a_autor, n_autor, genero, año, isbn, fecha = datos
                                        hoja.cell(row=renglon, column=1).value=clave
                                        hoja.cell(row=renglon, column=2).value=titulo
                                        hoja.cell(row=renglon, column=3).value=a_autor
                                        hoja.cell(row=renglon, column=4).value=n_autor
                                        hoja.cell(row=renglon, column=5).value=genero
                                        hoja.cell(row=renglon, column=6).value=año
                                        hoja.cell(row=renglon, column=7).value=isbn
                                        hoja.cell(row=renglon, column=8).value=fecha
                                        renglon += 1
                                    nombre_xl.save('Reporte_completo.xlsx')
                                elif exportar==3:
                                    print("Saliendo")
                                    break
                        elif reportes==2:
                            while True:  #ESTE WHILE NO SE SI TIENE BREAK XD
                                try:
                                    with sqlite3.connect(nombre_bd) as conn:
                                        cursor = conn.cursor()
                                        cursor.execute("SELECT * FROM autor")
                                        registros = cursor.fetchall()

                                        if registros:
                                            print('Los autores disponibles son')
                                            print("Claves\tApellido\tNombre")
                                            print("*" * 30)
                                            for clave, apellido, nombre in registros:
                                                print(f"{clave:^6}\t{apellido}\t{nombre}")
                                        else:
                                            print("No se encontraron Autores disponibles")
                                            break
                                except Error as e:
                                    print (e)
                                except Exception:
                                    print(f"Se produjo el siguiente error: {sys.exc_info()[0]}")
                                finally:
                                    conn.close()
                                autor_busqueda=int(input('Ingrese la clave del autor: '))
                                try:
                                    with sqlite3.connect(nombre_bd) as conn:
                                        cursor = conn.cursor()
                                        valores = {"autor_id":autor_busqueda}
                                        cursor.execute("SELECT libros.libro_id, libros.titulo, autor.apellido, autor.nombre, genero.nombre, libros.año, libros.isbn, libros.fecha \
                                                   FROM libros\
                                                   INNER JOIN autor ON libros.autor_id = autor.autor_id\
                                                   INNER JOIN genero ON libros.genero_id = genero.genero_id\
                                                   WHERE autor.autor_id = :autor_id;", valores)
                                        registro_libro = cursor.fetchall()

                                        if registro_libro:
                                            print("Clave\t\tTitulo\t\tApellido Autor\t\tNombre Autor\t\tGenero\t\tAño de Publicación\t\tISBN\t\tFecha de Adquisición")
                                            for clave, titulo, a_autor, n_autor, genero, año, isbn, fecha in registro_libro:
                                                print(f"{clave}\t{titulo}\t{a_autor}\t{n_autor}\t{genero}\t{año}\t{isbn}\t{fecha}")
                                                obras_autor.append((clave, titulo, a_autor, n_autor, genero, año, isbn, fecha))
                                                reporte_filtro.append((clave, titulo, a_autor, n_autor, genero, año, isbn, fecha))
                                            autor_aut=(f"Reporte_Autor_{a_autor}.csv")
                                        
                                            while True:
                                                exportar=int(input('Si quiere exportar seleccione: \n1: CSV \n2: Excel \n3:No exportar y salir\n'))
                                                if exportar==1:
                                                    with open(autor_aut,'w',newline='') as archivo:
                                                        grabador = csv.writer(archivo)
                                                        grabador.writerow(("Clave", "Titulo", "Apellido Autor", "Nombre Autor" "Genero", "Año", "ISBN", "Fecha de adquisicion"))
                                                        grabador.writerows([(clave, titulo, a_autor, n_autor, genero, año, isbn, fecha) for clave, titulo, a_autor, n_autor, genero, año, isbn, fecha in obras_autor])
                                                    print(f"El archivo fue guardado con el nombre: {autor_aut}")
                                                
                                                elif exportar==2:
                                                    nombre_xl = xl.Workbook()
                                                    hoja = nombre_xl.active
                                                    hoja = nombre_xl["Sheet"] 
                                                    hoja.title = "Reporte por Autor"                                   
                                                    hoja["A1"].value = "Clave"
                                                    hoja["B1"].value = "Titulo"
                                                    hoja["C1"].value = "Apellido Autor"
                                                    hoja["D1"].value = "Nombre Autor"
                                                    hoja["E1"].value = "Genero"
                                                    hoja["F1"].value = "Año"
                                                    hoja["G1"].value = "ISBN"
                                                    hoja["H1"].value = "Fecha de adquisicion"
                                                    
                                                    renglon = 2
                                                    for clave, titulo, a_autor, n_autor, genero, año, isbn, fecha in obras_autor:
                                                        hoja.cell(row=renglon, column=1).value=clave
                                                        hoja.cell(row=renglon, column=2).value=titulo
                                                        hoja.cell(row=renglon, column=3).value=a_autor
                                                        hoja.cell(row=renglon, column=4).value=n_autor
                                                        hoja.cell(row=renglon, column=5).value=genero
                                                        hoja.cell(row=renglon, column=6).value=año
                                                        hoja.cell(row=renglon, column=7).value=isbn
                                                        hoja.cell(row=renglon, column=8).value=fecha
                                                        renglon += 1
                                                    nombre_xl.save(f"Reporte_{a_autor}.xlsx")
                                                    
                                                elif exportar==3:
                                                    print("Saliendo")
                                                    break
                                        else:
                                            print(f"No se encontró un proyecto asociado con la clave {autor_busqueda}")
                                            continue
                                except Error as e:
                                    print (e)
                                    continue
                                except:
                                    print(f"Se produjo el siguiente error: {sys.exc_info()[0]}")
                                    continue
                                finally:
                                    conn.close()
                                    break
                                
                        elif reportes==3:  #ESTE TAMPOCO UWU
                            while True:
                                try:
                                    with sqlite3.connect(nombre_bd) as conn:
                                        cursor = conn.cursor()
                                        cursor.execute("SELECT * FROM genero")
                                        registros = cursor.fetchall()

                                        if registros:
                                            print('Los generos disponibles son')
                                            print("Claves\tNombre")
                                            print("*" * 30)
                                            for clave, nombre in registros:
                                                print(f"{clave:^6}\t{nombre}")
                                        else:
                                            print("No se encontraron Generos disponibles")
                                except Error as e:
                                    print (e)
                                except Exception:
                                    print(f"Se produjo el siguiente error: {sys.exc_info()[0]}")
                                finally:
                                    conn.close()
                                genero_busqueda=int(input('Ingrese la clave del genero: '))
                                try:
                                    with sqlite3.connect(nombre_bd) as conn:
                                        cursor = conn.cursor()
                                        valores = {"genero_id":genero_busqueda}
                                        cursor.execute("SELECT libros.libro_id, libros.titulo, autor.apellido, autor.nombre, genero.nombre, libros.año, libros.isbn, libros.fecha\
                                                   FROM libros \
                                                   INNER JOIN autor ON libros.autor_id = autor.autor_id\
                                                   INNER JOIN genero ON libros.genero_id = genero.genero_id\
                                                   WHERE genero.genero_id = :genero_id", valores)
                                        registro_libro = cursor.fetchall()

                                        if registro_libro:
                                            print("Clave\t\tTitulo\t\tApellido Autor\t\tNombre Autor\t\tGenero\t\tAño de Publicación\t\tISBN\t\tFecha de Adquisición")
                                            for clave, titulo, a_autor, n_autor, genero, año, isbn, fecha in registro_libro:
                                                print(f"{clave}\t{titulo}\t{a_autor}\t{n_autor}\t{genero}\t{año}\t{isbn}\t{fecha}")
                                                obras_genero.append((clave, titulo, a_autor, n_autor, genero, año, isbn, fecha))
                                                reporte_filtro.append((clave, titulo, a_autor, n_autor, genero, año, isbn, fecha))
                                            generos_aut=(f"Reporte_Genero_{genero}.csv")

                                            while True:
                                                exportar=int(input('Si quiere exportar seleccione: \n1: CSV \n2: Excel \n3:No exportar y salir\n'))
                                                if exportar==1:
                                                    with open(generos_aut,'w',newline='') as archivo:
                                                        grabador = csv.writer(archivo)
                                                        grabador.writerow(("Clave", "Titulo", "Apellido Autor", "Nombre Autor" "Genero", "Año", "ISBN", "Fecha de adquisicion"))
                                                        grabador.writerows([(clave, titulo, a_autor, n_autor, genero, año, isbn, fecha) for clave, titulo, a_autor, n_autor, genero, año, isbn, fecha in obras_genero])
                                                    print(f"El archivo fue guardado con el nombre: {generos_aut}")
                                                    
                                                elif exportar==2:
                                                    nombre_xl = xl.Workbook()
                                                    hoja = nombre_xl["Sheet"] 
                                                    hoja.title = "Reporte por Genero"
                                                    hoja["A1"].value = "Clave"
                                                    hoja["B1"].value = "Titulo"
                                                    hoja["C1"].value = "Apellido Autor"
                                                    hoja["D1"].value = "Nombre Autor"
                                                    hoja["E1"].value = "Genero"
                                                    hoja["F1"].value = "Año"
                                                    hoja["G1"].value = "ISBN"
                                                    hoja["H1"].value = "Fecha de adquisicion"
                                                    
                                                    renglon = 2
                                                    for clave, titulo, a_autor, n_autor, genero, año, isbn, fecha in obras_genero:
                                                        hoja.cell(row=renglon, column=1).value=clave
                                                        hoja.cell(row=renglon, column=2).value=titulo
                                                        hoja.cell(row=renglon, column=3).value=a_autor
                                                        hoja.cell(row=renglon, column=4).value=n_autor
                                                        hoja.cell(row=renglon, column=5).value=genero
                                                        hoja.cell(row=renglon, column=6).value=año
                                                        hoja.cell(row=renglon, column=7).value=isbn
                                                        hoja.cell(row=renglon, column=8).value=fecha
                                                        renglon += 1
                                                    nombre_xl.save(f"Reporte_{genero}.xlsx")
                                                    
                                                elif exportar==3:
                                                    print("Saliendo")
                                                    break
                                except Error as e:
                                    print (e)
                                    continue
                                except:
                                    print(f"Se produjo el siguiente error: {sys.exc_info()[0]}")
                                    continue
                                finally:
                                    conn.close()
                                    break

                        elif reportes==4:
                            while True:
                                try:
                                    with sqlite3.connect(nombre_bd) as conn:
                                        cursor = conn.cursor()
                                        cursor.execute("SELECT año FROM libros")
                                        registros = cursor.fetchall()

                                        if registros:
                                            print('Los años disponibles son')
                                            print("Año")
                                            print("*" * 30)
                                            for año in registros:
                                                print(f"{año}")
                                        else:
                                            print("No se encontraron Años disponibles")
                                except Error as e:
                                    print (e)
                                except Exception:
                                    print(f"Se produjo el siguiente error: {sys.exc_info()[0]}")
                                finally:
                                    conn.close()
                                año_busqueda=int(input('Ingrese el año de los libros a buscar: '))
                                try:
                                    with sqlite3.connect(nombre_bd) as conn:
                                        cursor = conn.cursor()
                                        valores = {"año":año_busqueda}
                                        cursor.execute("SELECT libros.libro_id, libros.titulo, autor.apellido, autor.nombre, genero.nombre, libros.año, libros.isbn, libros.fecha \
                                                   FROM libros \
                                                   INNER JOIN autor ON libros.autor_id = autor.autor_id\
                                                   INNER JOIN genero ON libros.genero_id = genero.genero_id\
                                                   WHERE año = :año", valores)
                                        registro_libro = cursor.fetchall()

                                        if registro_libro:
                                            print("Clave\t\tTitulo\t\tApellido Autor\t\tNombre Autor\t\tGenero\t\tAño de Publicación\t\tISBN\t\tFecha de Adquisición")
                                            for clave, titulo, a_autor, n_autor, genero, año, isbn, fecha in registro_libro:
                                                print(f"{clave}\t{titulo}\t{a_autor}\t{n_autor}\t{genero}\t{año}\t{isbn}\t{fecha}")
                                                obras_año.append((clave, titulo, a_autor, n_autor, genero, año, isbn, fecha))
                                                reporte_filtro.append((clave, titulo, a_autor, n_autor, genero, año, isbn, fecha))
                                            años_aut=(f"Reporte_Año_{año}.csv")

                                            while True:
                                                exportar=int(input('Si quiere exportar seleccione: \n1: CSV \n2: Excel \n3:No exportar y salir\n'))
                                                if exportar==1:
                                                    with open(años_aut,'w',newline='') as archivo:
                                                        grabador = csv.writer(archivo)
                                                        grabador.writerow(("Clave", "Titulo", "Apellido Autor", "Nombre Autor" "Genero", "Año", "ISBN", "Fecha de adquisicion"))
                                                        grabador.writerows([(clave, titulo, a_autor, n_autor, genero, año, isbn, fecha) for clave, titulo, a_autor, n_autor, genero, año, isbn, fecha in obras_año])
                                                    print(f"El archivo fue guardado con el nombre: {años_aut}")
                                                    
                                                elif exportar==2:
                                                    nombre_xl = xl.Workbook()
                                                    hoja = nombre_xl["Sheet"] 
                                                    hoja.title = "Reporte por Año"
                                                    hoja["A1"].value = "Clave"
                                                    hoja["B1"].value = "Titulo"
                                                    hoja["C1"].value = "Apellido Autor"
                                                    hoja["D1"].value = "Nombre Autor"
                                                    hoja["E1"].value = "Genero"
                                                    hoja["F1"].value = "Año"
                                                    hoja["G1"].value = "ISBN"
                                                    hoja["H1"].value = "Fecha de adquisicion"
                                                    
                                                    renglon = 2
                                                    for clave, titulo, a_autor, n_autor, genero, año, isbn, fecha in obras_año:
                                                        hoja.cell(row=renglon, column=1).value=clave
                                                        hoja.cell(row=renglon, column=2).value=titulo
                                                        hoja.cell(row=renglon, column=3).value=a_autor
                                                        hoja.cell(row=renglon, column=4).value=n_autor
                                                        hoja.cell(row=renglon, column=5).value=genero
                                                        hoja.cell(row=renglon, column=6).value=año
                                                        hoja.cell(row=renglon, column=7).value=isbn
                                                        hoja.cell(row=renglon, column=8).value=fecha
                                                        renglon += 1
                                                    nombre_xl.save(f"Reporte_{año}.xlsx")
                                                    
                                                elif exportar==3:
                                                    print("Saliendo")
                                                    break
                                except Error as e:
                                    print (e)
                                    continue
                                except:
                                    print(f"Se produjo el siguiente error: {sys.exc_info()[0]}")
                                    continue
                                finally:
                                    conn.close()
                                    break
                        elif reportes==5:
                            while True:
                                exportar=int(input('Si quiere exportar seleccione: \n1: CSV \n2: Excel \n3:No exportar y salir\n'))
                                if exportar==1:
                                    with open(nombre_csv,'w',newline='') as archivo:
                                        grabador = csv.writer(archivo)
                                        grabador.writerow(("Clave", "Titulo", "Apellido Autor", "Nombre Autor" "Genero", "Año", "ISBN", "Fecha de adquisicion"))
                                        grabador.writerows([(clave, titulo, a_autor, n_autor, genero, año, isbn, fecha) for clave, titulo, a_autor, n_autor, genero, año, isbn, fecha in reporte_filtro])
                                    print(f"El archivo fue guardado con el nombre: {nombre_csv}")
                                elif exportar==2:
                                    nombre_xl = xl.Workbook()
                                    hoja = nombre_xl["Sheet"] 
                                    hoja.title = "Reporte Filtrado"
                                    hoja["A1"].value = "Clave"
                                    hoja["B1"].value = "Titulo"
                                    hoja["C1"].value = "Apellido Autor"
                                    hoja["D1"].value = "Nombre Autor"
                                    hoja["E1"].value = "Genero"
                                    hoja["F1"].value = "Año"
                                    hoja["G1"].value = "ISBN"
                                    hoja["H1"].value = "Fecha de adquisicion"
                                    
                                    renglon = 2
                                    for clave, datos in libros.items():
                                        titulo, a_autor, n_autor, genero, año, isbn, fecha = datos
                                        hoja.cell(row=renglon, column=1).value=clave
                                        hoja.cell(row=renglon, column=2).value=titulo
                                        hoja.cell(row=renglon, column=3).value=a_autor
                                        hoja.cell(row=renglon, column=4).value=n_autor
                                        hoja.cell(row=renglon, column=5).value=genero
                                        hoja.cell(row=renglon, column=6).value=año
                                        hoja.cell(row=renglon, column=7).value=isbn
                                        hoja.cell(row=renglon, column=8).value=fecha
                                        renglon += 1
                                    nombre_xl.save('libros_filtro.xlsx')
                                    
                                elif exportar==3:
                                    print("Saliendo")
                                    break
                            break
                        else:
                            print('Opcion no valida')
                            continue
                elif menu==3:
                    break
                else:
                    print('Opcion no valida')
                    continue
        elif opcion==3:
            while True:
                ap_autor=input('Ingresa el apellido del autor:\n').upper()
                if ap_autor=='':
                    print('No se permite dejar vacios. Intente de nuevo')
                    continue
                else:
                    break
            while True:
                nom_autor=input('Ingresa el nombre del autor a registrar:\n').upper()
                if nom_autor=='':
                    print('No se permite dejar vacios. Intente de nuevo')
                    continue
                else:
                    break
            try:
                with sqlite3.connect(nombre_bd) as conn:
                    cursor = conn.cursor()
                    valores = (ap_autor, nom_autor)
                    cursor.execute("INSERT INTO autor (apellido, nombre) VALUES(?,?)", valores)
                    print(f"La clave asignada fue {cursor.lastrowid}")
            except Error as e:
                print (e)
            except Exception:
                print(f"Se produjo el siguiente error: {sys.exc_info()[0]}")
            finally:
                conn.close()
        
        elif opcion==4:
            while True:
                nom_genero=input('Ingresa el nombre del genero:\n').upper()
                if nom_genero=='':
                    print('No se permite dejar vacios. Intente de nuevo')
                    continue
                else:
                    break
            try:
                with sqlite3.connect(nombre_bd) as conn:
                    cursor = conn.cursor()
                    valores = (nom_genero,)
                    cursor.execute("INSERT INTO genero (nombre) VALUES(?)", valores)
                    print(f"La clave asignada fue {cursor.lastrowid}")
            except Error as e:
                print (e)
            except Exception:
                print(f"Se produjo el siguiente error: {sys.exc_info()[0]}")
            finally:
                conn.close()

        elif opcion==5:
            with open(nombre_csv,"w", newline="") as archivo:
                grabador = csv.writer(archivo)
                grabador.writerow(("Clave", "Titulo", "Apellido Autor", "Nombre Autor" "Genero", "Año", "ISBN", "Fecha de adquisicion"))
                grabador.writerows([(clave, datos[0], datos[1], datos[2], datos[3], datos[4], datos[5], datos[6]) for clave, datos in libros.items()])
            print(f"El archivo fue guardado con el nombre: {nombre_csv}")
            break
        else:
            print('Opcion no valida')
else:
    print('Se encontro el archivo y se procede a usarlo')
    while True:
        try:
            try:
                with sqlite3.connect(nombre_bd) as conn:
                    cursor = conn.cursor()
                    cursor.execute("SELECT * FROM autor")
                    registros = cursor.fetchall()

                    if registros:
                        print('Los autores disponibles son')
                        print("Claves\tApellido\tNombre")
                        print("*" * 30)
                        for clave, apellido, nombre in registros:
                            print(f"{clave:^6}\t{apellido}\t{nombre}")
                    else:
                        print("No se encontraron Autores disponibles, intenta ingresar algunos antes de ingresar libros")
            except Error as e:
                print (e)
            except Exception:
                print(f"Se produjo el siguiente error: {sys.exc_info()[0]}")
            finally:
                conn.close()

            try:
                with sqlite3.connect(nombre_bd) as conn:
                    cursor = conn.cursor()
                    cursor.execute("SELECT * FROM genero")
                    registros = cursor.fetchall()

                    if registros:
                        print('Los generos disponibles son')
                        print("Claves\tNombre")
                        print("*" * 30)
                        for clave, nombre in registros:
                            print(f"{clave:^6}\t{nombre}")
                    else:
                        print("No se encontraron Generos disponibles, intenta ingresar algunos antes de ingresar libros")
            except Error as e:
                print (e)
            except Exception:
                print(f"Se produjo el siguiente error: {sys.exc_info()[0]}")
            finally:
                conn.close()
        except Error as e:
            print (e)
        except Exception:
            print(f"Se produjo el siguiente error: {sys.exc_info()[0]}")
        finally:
            conn.close()

        opcion=int(input(f'Seleccione alguna de las siguientes opciones:\n 1:Subir libro\n 2:Consultas y Reportes\n 3:Registrar Autor\n 4:Registrar Genero\n 5:Salir\n'))

        if opcion==1:
            clave=max(libros, default=0)+1
            while True:
                titulo=input('Ingrese el titulo del libro: ').upper()
                if titulo=='':
                    print('No se permite dejar vacios. Intente de nuevo ')
                    continue
                else:
                    break
            while True:
                try:
                    with sqlite3.connect(nombre_bd) as conn:
                        cursor = conn.cursor()
                        cursor.execute("SELECT * FROM autor")
                        registros = cursor.fetchall()

                        if registros:
                            print('Los autores disponibles son')
                            print("Claves\tApellido\tNombre")
                            print("*" * 30)
                            for clave, apellido, nombre in registros:
                                print(f"{clave:^6}\t{apellido}\t{nombre}")
                        else:
                            print("No se encontraron Autores disponibles, intenta ingresar algunos antes de ingresar libros")
                            nom_autor=input('Ingresa el nombre del autor a registrar:\n').upper()
                            if nom_autor=='':
                                print('No se permite dejar vacios. Intente de nuevo')
                                continue
                            ap_autor=input('Ingresa el apellido del autor:\n').upper()
                            if ap_autor=='':
                                print('No se permite dejar vacios. Intente de nuevo')
                                continue
                            else:
                                try:
                                    with sqlite3.connect(nombre_bd) as conn:
                                        cursor = conn.cursor()
                                        valores = (ap_autor, nom_autor)
                                        cursor.execute("INSERT INTO autor (apellido, nombre) VALUES(?,?)", valores)
                                        print(f"La clave asignada fue {cursor.lastrowid}")
                                except Error as e:
                                    print (e)
                                except Exception:
                                    print(f"Se produjo el siguiente error: {sys.exc_info()[0]}")
                                finally:
                                    conn.close()
                except Error as e:
                    print (e)
                except Exception:
                    print(f"Se produjo el siguiente error: {sys.exc_info()[0]}")
                finally:
                    conn.close()
                try:
                    id_autor=int(input('Ingrese la clave del autor del libro: '))
                except ValueError:
                    print('No se permite dejar vacios. Intente de nuevo ')
                    continue
                else:
                    break
            while True:
                try:
                    with sqlite3.connect(nombre_bd) as conn:
                        cursor = conn.cursor()
                        cursor.execute("SELECT * FROM genero")
                        registros = cursor.fetchall()

                        if registros:
                            print('Los generos disponibles son')
                            print("Claves\tNombre")
                            print("*" * 30)
                            for clave, nombre in registros:
                                print(f"{clave:^6}\t{nombre}")
                        else:
                            print("No se encontraron Generos disponibles, intenta ingresar algunos antes de ingresar libros")
                            while True:
                                nom_genero=input('Ingresa el nombre del genero:\n').upper()
                                if nom_genero=='':
                                    print('No se permite dejar vacios. Intente de nuevo')
                                    continue
                                else:
                                    break
                            try:
                                with sqlite3.connect(nombre_bd) as conn:
                                    cursor = conn.cursor()
                                    valores = (nom_genero,)
                                    cursor.execute("INSERT INTO genero (nombre) VALUES(?)", valores)
                                    print(f"La clave asignada fue {cursor.lastrowid}")
                            except Error as e:
                                print (e)
                            except Exception:
                                print(f"Se produjo el siguiente error: {sys.exc_info()[0]}")
                            finally:
                                conn.close()
                except Error as e:
                    print (e)
                except:
                    print(f"Se produjo el siguiente error: {sys.exc_info()[0]}")
                finally:
                    conn.close()
                try:
                    id_genero=int(input('Ingrese la clave del genero del libro: '))
                except ValueError:
                    print('No se permite dejar vacios. Intente de nuevo ')
                    continue
                else:
                    break
            while True:
                try:
                    año=int(input('Ingrese el año en que fue publicado el libro: '))
                except ValueError:
                    print('No son validas las letras en el año ')
                    continue
                else:
                    if len(str(año)) > 4:
                        print('No existe ese. Intente de nuevo ')
                        continue
                    else:
                        break
            while True:
                try:      
                    isbn=int(input('Ingrese el ISBN del libro: '))
                except ValueError:
                    print('No son validas las letras en el ISBN')
                    continue
                else:
                    if isbn=='':
                        print('No se permite dejar vacios. Intente de nuevo')
                        continue
                    else:
                        break
            while True:
                fecha_adquisicion=input('Ingrese la fecha en que fue adquirido el libro (dd/mm/aaaa): ')
                if fecha_adquisicion=='':
                    print('No se permite dejar vacios. Intente de nuevo ')
                    continue
                if (not bool(re.match(f_fecha, fecha_adquisicion))):
                    print("La fecha inicial no tiene el formato (dd/mm/aaaa)")
                    continue
                else:
                    try:
                        fecha_adq= datetime.datetime.strptime(fecha_adquisicion, "%d/%m/%Y").date()
                    except:
                        print('La fecha ingresada no existe, ingresa una nueva')
                        continue
                    fecha = (fecha_adq.strftime("%d/%m/%Y"))
                    break
            try:
                with sqlite3.connect(nombre_bd) as conn:
                    cursor = conn.cursor()
                    libro=(titulo,id_autor,id_genero,año,isbn,fecha)
                    cursor.execute("INSERT INTO libros (titulo, autor_id, genero_id, año, isbn, fecha) VALUES(?,?,?,?,?,?)", libro)
                    print("Registro agregado exitosamente")
            except Error as e:
                print (e)
            except:
                print(f"Se produjo el siguiente error: {sys.exc_info()[0]}")
            finally:
                conn.close()
            try:
                with sqlite3.connect(nombre_bd) as conn:
                    cursor = conn.cursor()
                    cursor.execute("SELECT libros.libro_id, libros.titulo, autor.apellido, autor.nombre, genero.nombre, libros.año, libros.isbn, libros.fecha \
                                    FROM libros \
                                    INNER JOIN autor ON libros.autor_id = autor.autor_id\
                                    INNER JOIN genero ON libros.genero_id = genero.genero_id\
                                    ")
                    registro_libro = cursor.fetchall()

                    if registro_libro:
                        print("\t\tClave\t\tTitulo\t\tApellido Autor\t\tNombre Autor\t\tGenero\t\tAño de Publicación\t\tISBN\t\tFecha de Adquisición")
                        for clave, titulo, a_autor, n_autor, genero, año, isbn, fecha in registro_libro:
                            print(f"{clave}\t{titulo}\t{a_autor}\t{n_autor}\t{genero}\t{año}\t{isbn}\t{fecha}")
                            libros[clave]=[titulo,a_autor,n_autor,genero,año,isbn,fecha]
                            libros_clave.append((clave, titulo, a_autor, n_autor, genero, int(año), int(isbn), fecha))
                    else:
                        print('No se encontro ningun libro')
                        continue
            except Error as e:
                print (e)
            except:
                print(f"Se produjo el siguiente error: {sys.exc_info()[0]}")
            finally:
                conn.close()

        elif opcion==2:
            while True:
                menu=int(input(f'Seleccione alguna de las siguientes opciones:\n1:Consulta de titulo\n2:Reportes\n3:Volver al menu\n'))
                if menu==1:
                    while True:
                        consulta=int(input(f'Seleccione alguna de las siguientes Consultas:\n1:Por titulo\n2:Por ISBN\n3:Salir\n'))
                        if consulta==1:
                            try:
                                with sqlite3.connect(nombre_bd) as conn:
                                    cursor = conn.cursor()
                                    cursor.execute("SELECT titulo FROM libros")
                                    titulos_registro = cursor.fetchall()

                                    if titulos_registro:
                                        print("Titulos")
                                        for titulo in titulos_registro:
                                            print(f"{titulo}")
                                    else:
                                        print('No se encontro ningun libro')
                                        break
                            except Error as e:
                                print (e)
                                continue
                            except:
                                print(f"Se produjo el siguiente error: {sys.exc_info()[0]}")
                                continue
                            finally:
                                conn.close()
                            consulta_titulo=input('Ingrese el nombre del titulo a buscar: ').upper()
                            try:
                                with sqlite3.connect(nombre_bd) as conn:
                                    cursor = conn.cursor()
                                    valores = {"titulo":consulta_titulo}
                                    cursor.execute("SELECT libros.libro_id, libros.titulo, autor.apellido, autor.nombre, genero.nombre, libros.año, libros.isbn, libros.fecha \
                                                   FROM libros \
                                                   INNER JOIN autor ON libros.autor_id = autor.autor_id\
                                                   INNER JOIN genero ON libros.genero_id = genero.genero_id\
                                                   WHERE titulo = :titulo", valores)
                                    registro_libro = cursor.fetchall()

                                    if registro_libro:
                                        print("Clave\t\tTitulo\t\tApellido Autor\t\tNombre Autor\t\tGenero\t\tAño de Publicación\t\tISBN\t\tFecha de Adquisición")
                                        for clave, titulo, a_autor, n_autor, genero, año, isbn, fecha in registro_libro:
                                            print(f"{clave}\t{titulo}\t{a_autor}\t{n_autor}\t{genero}\t{año}\t{isbn}\t{fecha}")
                                    else:
                                        print('No se encontro ningun libro con ese nombre')
                                        continue
                            except Error as e:
                                print (e)

                            except:
                                print(f"Se produjo el siguiente error: {sys.exc_info()[0]}")
                            finally:
                                conn.close()

                        elif consulta==2:
                            try:
                                with sqlite3.connect(nombre_bd) as conn:
                                    cursor = conn.cursor()
                                    cursor.execute("SELECT isbn FROM libros")
                                    isbn_registro = cursor.fetchall()

                                    if isbn_registro:
                                        print("ISBN")
                                        for isbn in isbn_registro:
                                            print(f"{isbn}")
                                    else:
                                        print('No se encontro ningun libro')
                                        break
                            except Error as e:
                                print (e)
                                continue
                            except:
                                print(f"Se produjo el siguiente error: {sys.exc_info()[0]}")
                                continue
                            finally:
                                conn.close()
                            consulta_isbn=int(input('Ingrese el isbn a buscar: '))
                            try:
                                with sqlite3.connect(nombre_bd) as conn:
                                    cursor = conn.cursor()
                                    valores = {"isbn":consulta_isbn}
                                    cursor.execute("SELECT libros.libro_id, libros.titulo, autor.apellido, autor.nombre, genero.nombre, libros.año, libros.isbn, libros.fecha \
                                                   FROM libros \
                                                   INNER JOIN autor ON libros.autor_id = autor.autor_id\
                                                   INNER JOIN genero ON libros.genero_id = genero.genero_id\
                                                   WHERE isbn = :isbn", valores)
                                    registro_libro = cursor.fetchall()

                                    if registro_libro:
                                        print("Clave\t\tTitulo\t\tApellido Autor\t\tNombre Autor\t\tGenero\t\tAño de Publicación\t\tISBN\t\tFecha de Adquisición")
                                        for clave, titulo, a_autor, n_autor, genero, año, isbn, fecha in registro_libro:
                                            print(f"{clave}\t{titulo}\t{a_autor}\t{n_autor}\t{genero}\t{año}\t{isbn}\t{fecha}")
                                    else:
                                        print('No se encontro ningun libro con ese isbn')
                                        continue
                            except Error as e:
                                print (e)
                                continue
                            except:
                                print(f"Se produjo el siguiente error: {sys.exc_info()[0]}")
                                continue
                            finally:
                                conn.close()
                        elif consulta==3:
                            break
                        else:
                            print('Opcion no valida')
                            continue

                elif menu==2:
                    while True:
                        reportes=int(input(f"Seleccione por que medio desea realizar los reportes:\n1:Ver catalogo completo\n2:Por autor\n3:Genero\n4:Año publicacion\n5:Volver al menu consultas y reportes\n"))
                        if reportes==1:
                            diccionario = libros.values()
                            print("\n** Catálogo completo ** ")
                            print("Titulo\t\t\tApellido Autor\t\t\tNombre Autor\t\t\tGenero\t\t\tAño de Publicación\t\t\tISBN\t\t\tFecha de Adquisición")
                            print("*" * 120)
                            for titulo, a_autor, n_autor, genero, año_publicacion, isbn, fecha in diccionario:
                                print(f"{titulo: <15} | {a_autor: <15} | {n_autor: <15} | {genero: <15} | {año_publicacion: <15} | {isbn:<15} |  {fecha:<15}")
                        
                            while True:
                                exportar=int(input('Si quiere exportar seleccione: \n1: CSV \n2: Excel \n3:No exportar y salir\n'))
                                if exportar==1:
                                    with open(nombre_csv,'w',newline='') as archivo:
                                        grabador = csv.writer(archivo)
                                        grabador.writerow(("Clave", "Titulo", "Apellido Autor", "Nombre Autor" "Genero", "Año", "ISBN", "Fecha de adquisicion"))
                                        grabador.writerows([(clave, datos[0], datos[1], datos[2], datos[3], datos[4], datos[5], datos[6]) for clave, datos in libros.items()])
                                    print(f"El archivo fue guardado con el nombre: {nombre_csv}")
                                elif exportar==2:
                                    nombre_xl = xl.Workbook()
                                    
                                    hoja = nombre_xl["Sheet"] 
                                    hoja.title = "Reporte Completo"
                                    
                                    hoja["A1"].value = "Clave"
                                    hoja["B1"].value = "Titulo"
                                    hoja["C1"].value = "Apellido Autor"
                                    hoja["D1"].value = "Nombre Autor"
                                    hoja["E1"].value = "Genero"
                                    hoja["F1"].value = "Año"
                                    hoja["G1"].value = "ISBN"
                                    hoja["H1"].value = "Fecha de adquisicion"
                                    
                                    renglon = 2
                                    for clave, datos in libros.items():
                                        titulo, a_autor, n_autor, genero, año, isbn, fecha = datos
                                        hoja.cell(row=renglon, column=1).value=clave
                                        hoja.cell(row=renglon, column=2).value=titulo
                                        hoja.cell(row=renglon, column=3).value=a_autor
                                        hoja.cell(row=renglon, column=4).value=n_autor
                                        hoja.cell(row=renglon, column=5).value=genero
                                        hoja.cell(row=renglon, column=6).value=año
                                        hoja.cell(row=renglon, column=7).value=isbn
                                        hoja.cell(row=renglon, column=8).value=fecha
                                        renglon += 1
                                    nombre_xl.save('Reporte_completo.xlsx')
                                elif exportar==3:
                                    print("Saliendo")
                                    break
                        elif reportes==2:
                            while True:  #ESTE WHILE NO SE SI TIENE BREAK XD
                                try:
                                    with sqlite3.connect(nombre_bd) as conn:
                                        cursor = conn.cursor()
                                        cursor.execute("SELECT * FROM autor")
                                        registros = cursor.fetchall()

                                        if registros:
                                            print('Los autores disponibles son')
                                            print("Claves\tApellido\tNombre")
                                            print("*" * 30)
                                            for clave, apellido, nombre in registros:
                                                print(f"{clave:^6}\t{apellido}\t{nombre}")
                                        else:
                                            print("No se encontraron Autores disponibles")
                                            break
                                except Error as e:
                                    print (e)
                                except Exception:
                                    print(f"Se produjo el siguiente error: {sys.exc_info()[0]}")
                                finally:
                                    conn.close()
                                autor_busqueda=int(input('Ingrese la clave del autor: '))
                                try:
                                    with sqlite3.connect(nombre_bd) as conn:
                                        cursor = conn.cursor()
                                        valores = {"autor_id":autor_busqueda}
                                        cursor.execute("SELECT libros.libro_id, libros.titulo, autor.apellido, autor.nombre, genero.nombre, libros.año, libros.isbn, libros.fecha \
                                                   FROM libros\
                                                   INNER JOIN autor ON libros.autor_id = autor.autor_id\
                                                   INNER JOIN genero ON libros.genero_id = genero.genero_id\
                                                   WHERE autor.autor_id = :autor_id;", valores)
                                        registro_libro = cursor.fetchall()

                                        if registro_libro:
                                            print("Clave\t\tTitulo\t\tApellido Autor\t\tNombre Autor\t\tGenero\t\tAño de Publicación\t\tISBN\t\tFecha de Adquisición")
                                            for clave, titulo, a_autor, n_autor, genero, año, isbn, fecha in registro_libro:
                                                print(f"{clave}\t{titulo}\t{a_autor}\t{n_autor}\t{genero}\t{año}\t{isbn}\t{fecha}")
                                                obras_autor.append((clave, titulo, a_autor, n_autor, genero, año, isbn, fecha))
                                                reporte_filtro.append((clave, titulo, a_autor, n_autor, genero, año, isbn, fecha))
                                            autor_aut=(f"Reporte_Autor_{a_autor}.csv")
                                        
                                            while True:
                                                exportar=int(input('Si quiere exportar seleccione: \n1: CSV \n2: Excel \n3:No exportar y salir\n'))
                                                if exportar==1:
                                                    with open(autor_aut,'w',newline='') as archivo:
                                                        grabador = csv.writer(archivo)
                                                        grabador.writerow(("Clave", "Titulo", "Apellido Autor", "Nombre Autor" "Genero", "Año", "ISBN", "Fecha de adquisicion"))
                                                        grabador.writerows([(clave, titulo, a_autor, n_autor, genero, año, isbn, fecha) for clave, titulo, a_autor, n_autor, genero, año, isbn, fecha in obras_autor])
                                                    print(f"El archivo fue guardado con el nombre: {autor_aut}")
                                                
                                                elif exportar==2:
                                                    nombre_xl = xl.Workbook()
                                                    hoja = nombre_xl.active
                                                    hoja = nombre_xl["Sheet"] 
                                                    hoja.title = "Reporte por Autor"                                   
                                                    hoja["A1"].value = "Clave"
                                                    hoja["B1"].value = "Titulo"
                                                    hoja["C1"].value = "Apellido Autor"
                                                    hoja["D1"].value = "Nombre Autor"
                                                    hoja["E1"].value = "Genero"
                                                    hoja["F1"].value = "Año"
                                                    hoja["G1"].value = "ISBN"
                                                    hoja["H1"].value = "Fecha de adquisicion"
                                                    
                                                    renglon = 2
                                                    for clave, titulo, a_autor, n_autor, genero, año, isbn, fecha in obras_autor:
                                                        hoja.cell(row=renglon, column=1).value=clave
                                                        hoja.cell(row=renglon, column=2).value=titulo
                                                        hoja.cell(row=renglon, column=3).value=a_autor
                                                        hoja.cell(row=renglon, column=4).value=n_autor
                                                        hoja.cell(row=renglon, column=5).value=genero
                                                        hoja.cell(row=renglon, column=6).value=año
                                                        hoja.cell(row=renglon, column=7).value=isbn
                                                        hoja.cell(row=renglon, column=8).value=fecha
                                                        renglon += 1
                                                    nombre_xl.save(f"Reporte_{a_autor}.xlsx")
                                                    
                                                elif exportar==3:
                                                    print("Saliendo")
                                                    break
                                        else:
                                            print(f"No se encontró un proyecto asociado con la clave {autor_busqueda}")
                                            continue
                                except Error as e:
                                    print (e)
                                    continue
                                except:
                                    print(f"Se produjo el siguiente error: {sys.exc_info()[0]}")
                                    continue
                                finally:
                                    conn.close()
                                    break
                                
                        elif reportes==3:  #ESTE TAMPOCO UWU
                            while True:
                                try:
                                    with sqlite3.connect(nombre_bd) as conn:
                                        cursor = conn.cursor()
                                        cursor.execute("SELECT * FROM genero")
                                        registros = cursor.fetchall()

                                        if registros:
                                            print('Los generos disponibles son')
                                            print("Claves\tNombre")
                                            print("*" * 30)
                                            for clave, nombre in registros:
                                                print(f"{clave:^6}\t{nombre}")
                                        else:
                                            print("No se encontraron Generos disponibles")
                                except Error as e:
                                    print (e)
                                except Exception:
                                    print(f"Se produjo el siguiente error: {sys.exc_info()[0]}")
                                finally:
                                    conn.close()
                                genero_busqueda=int(input('Ingrese la clave del genero: '))
                                try:
                                    with sqlite3.connect(nombre_bd) as conn:
                                        cursor = conn.cursor()
                                        valores = {"genero_id":genero_busqueda}
                                        cursor.execute("SELECT libros.libro_id, libros.titulo, autor.apellido, autor.nombre, genero.nombre, libros.año, libros.isbn, libros.fecha\
                                                   FROM libros \
                                                   INNER JOIN autor ON libros.autor_id = autor.autor_id\
                                                   INNER JOIN genero ON libros.genero_id = genero.genero_id\
                                                   WHERE genero.genero_id = :genero_id", valores)
                                        registro_libro = cursor.fetchall()

                                        if registro_libro:
                                            print("Clave\t\tTitulo\t\tApellido Autor\t\tNombre Autor\t\tGenero\t\tAño de Publicación\t\tISBN\t\tFecha de Adquisición")
                                            for clave, titulo, a_autor, n_autor, genero, año, isbn, fecha in registro_libro:
                                                print(f"{clave}\t{titulo}\t{a_autor}\t{n_autor}\t{genero}\t{año}\t{isbn}\t{fecha}")
                                                obras_genero.append((clave, titulo, a_autor, n_autor, genero, año, isbn, fecha))
                                                reporte_filtro.append((clave, titulo, a_autor, n_autor, genero, año, isbn, fecha))
                                            generos_aut=(f"Reporte_Genero_{genero}.csv")

                                            while True:
                                                exportar=int(input('Si quiere exportar seleccione: \n1: CSV \n2: Excel \n3:No exportar y salir\n'))
                                                if exportar==1:
                                                    with open(generos_aut,'w',newline='') as archivo:
                                                        grabador = csv.writer(archivo)
                                                        grabador.writerow(("Clave", "Titulo", "Apellido Autor", "Nombre Autor" "Genero", "Año", "ISBN", "Fecha de adquisicion"))
                                                        grabador.writerows([(clave, titulo, a_autor, n_autor, genero, año, isbn, fecha) for clave, titulo, a_autor, n_autor, genero, año, isbn, fecha in obras_genero])
                                                    print(f"El archivo fue guardado con el nombre: {generos_aut}")
                                                    
                                                elif exportar==2:
                                                    nombre_xl = xl.Workbook()
                                                    hoja = nombre_xl["Sheet"] 
                                                    hoja.title = "Reporte por Genero"
                                                    hoja["A1"].value = "Clave"
                                                    hoja["B1"].value = "Titulo"
                                                    hoja["C1"].value = "Apellido Autor"
                                                    hoja["D1"].value = "Nombre Autor"
                                                    hoja["E1"].value = "Genero"
                                                    hoja["F1"].value = "Año"
                                                    hoja["G1"].value = "ISBN"
                                                    hoja["H1"].value = "Fecha de adquisicion"
                                                    
                                                    renglon = 2
                                                    for clave, titulo, a_autor, n_autor, genero, año, isbn, fecha in obras_genero:
                                                        hoja.cell(row=renglon, column=1).value=clave
                                                        hoja.cell(row=renglon, column=2).value=titulo
                                                        hoja.cell(row=renglon, column=3).value=a_autor
                                                        hoja.cell(row=renglon, column=4).value=n_autor
                                                        hoja.cell(row=renglon, column=5).value=genero
                                                        hoja.cell(row=renglon, column=6).value=año
                                                        hoja.cell(row=renglon, column=7).value=isbn
                                                        hoja.cell(row=renglon, column=8).value=fecha
                                                        renglon += 1
                                                    nombre_xl.save(f"Reporte_{genero}.xlsx")
                                                    
                                                elif exportar==3:
                                                    print("Saliendo")
                                                    break
                                except Error as e:
                                    print (e)
                                    continue
                                except:
                                    print(f"Se produjo el siguiente error: {sys.exc_info()[0]}")
                                    continue
                                finally:
                                    conn.close()
                                    break

                        elif reportes==4:
                            while True:
                                try:
                                    with sqlite3.connect(nombre_bd) as conn:
                                        cursor = conn.cursor()
                                        cursor.execute("SELECT año FROM libros")
                                        registros = cursor.fetchall()

                                        if registros:
                                            print('Los años disponibles son')
                                            print("Año")
                                            print("*" * 30)
                                            for año in registros:
                                                print(f"{año}")
                                        else:
                                            print("No se encontraron Años disponibles")
                                except Error as e:
                                    print (e)
                                except Exception:
                                    print(f"Se produjo el siguiente error: {sys.exc_info()[0]}")
                                finally:
                                    conn.close()
                                año_busqueda=int(input('Ingrese el año de los libros a buscar: '))
                                try:
                                    with sqlite3.connect(nombre_bd) as conn:
                                        cursor = conn.cursor()
                                        valores = {"año":año_busqueda}
                                        cursor.execute("SELECT libros.libro_id, libros.titulo, autor.apellido, autor.nombre, genero.nombre, libros.año, libros.isbn, libros.fecha \
                                                   FROM libros \
                                                   INNER JOIN autor ON libros.autor_id = autor.autor_id\
                                                   INNER JOIN genero ON libros.genero_id = genero.genero_id\
                                                   WHERE año = :año", valores)
                                        registro_libro = cursor.fetchall()

                                        if registro_libro:
                                            print("Clave\t\tTitulo\t\tApellido Autor\t\tNombre Autor\t\tGenero\t\tAño de Publicación\t\tISBN\t\tFecha de Adquisición")
                                            for clave, titulo, a_autor, n_autor, genero, año, isbn, fecha in registro_libro:
                                                print(f"{clave}\t{titulo}\t{a_autor}\t{n_autor}\t{genero}\t{año}\t{isbn}\t{fecha}")
                                                obras_año.append((clave, titulo, a_autor, n_autor, genero, año, isbn, fecha))
                                                reporte_filtro.append((clave, titulo, a_autor, n_autor, genero, año, isbn, fecha))
                                            años_aut=(f"Reporte_Año_{año}.csv")

                                            while True:
                                                exportar=int(input('Si quiere exportar seleccione: \n1: CSV \n2: Excel \n3:No exportar y salir\n'))
                                                if exportar==1:
                                                    with open(años_aut,'w',newline='') as archivo:
                                                        grabador = csv.writer(archivo)
                                                        grabador.writerow(("Clave", "Titulo", "Apellido Autor", "Nombre Autor" "Genero", "Año", "ISBN", "Fecha de adquisicion"))
                                                        grabador.writerows([(clave, titulo, a_autor, n_autor, genero, año, isbn, fecha) for clave, titulo, a_autor, n_autor, genero, año, isbn, fecha in obras_año])
                                                    print(f"El archivo fue guardado con el nombre: {años_aut}")
                                                    
                                                elif exportar==2:
                                                    nombre_xl = xl.Workbook()
                                                    hoja = nombre_xl["Sheet"] 
                                                    hoja.title = "Reporte por Año"
                                                    hoja["A1"].value = "Clave"
                                                    hoja["B1"].value = "Titulo"
                                                    hoja["C1"].value = "Apellido Autor"
                                                    hoja["D1"].value = "Nombre Autor"
                                                    hoja["E1"].value = "Genero"
                                                    hoja["F1"].value = "Año"
                                                    hoja["G1"].value = "ISBN"
                                                    hoja["H1"].value = "Fecha de adquisicion"
                                                    
                                                    renglon = 2
                                                    for clave, titulo, a_autor, n_autor, genero, año, isbn, fecha in obras_año:
                                                        hoja.cell(row=renglon, column=1).value=clave
                                                        hoja.cell(row=renglon, column=2).value=titulo
                                                        hoja.cell(row=renglon, column=3).value=a_autor
                                                        hoja.cell(row=renglon, column=4).value=n_autor
                                                        hoja.cell(row=renglon, column=5).value=genero
                                                        hoja.cell(row=renglon, column=6).value=año
                                                        hoja.cell(row=renglon, column=7).value=isbn
                                                        hoja.cell(row=renglon, column=8).value=fecha
                                                        renglon += 1
                                                    nombre_xl.save(f"Reporte_{año}.xlsx")
                                                    
                                                elif exportar==3:
                                                    print("Saliendo")
                                                    break
                                except Error as e:
                                    print (e)
                                    continue
                                except:
                                    print(f"Se produjo el siguiente error: {sys.exc_info()[0]}")
                                    continue
                                finally:
                                    conn.close()
                                    break
                        elif reportes==5:
                            while True:
                                exportar=int(input('Si quiere exportar seleccione: \n1: CSV \n2: Excel \n3:No exportar y salir\n'))
                                if exportar==1:
                                    with open(nombre_csv,'w',newline='') as archivo:
                                        grabador = csv.writer(archivo)
                                        grabador.writerow(("Clave", "Titulo", "Apellido Autor", "Nombre Autor" "Genero", "Año", "ISBN", "Fecha de adquisicion"))
                                        grabador.writerows([(clave, titulo, a_autor, n_autor, genero, año, isbn, fecha) for clave, titulo, a_autor, n_autor, genero, año, isbn, fecha in reporte_filtro])
                                    print(f"El archivo fue guardado con el nombre: {nombre_csv}")
                                elif exportar==2:
                                    nombre_xl = xl.Workbook()
                                    hoja = nombre_xl["Sheet"] 
                                    hoja.title = "Reporte Filtrado"
                                    hoja["A1"].value = "Clave"
                                    hoja["B1"].value = "Titulo"
                                    hoja["C1"].value = "Apellido Autor"
                                    hoja["D1"].value = "Nombre Autor"
                                    hoja["E1"].value = "Genero"
                                    hoja["F1"].value = "Año"
                                    hoja["G1"].value = "ISBN"
                                    hoja["H1"].value = "Fecha de adquisicion"
                                    
                                    renglon = 2
                                    for clave, datos in libros.items():
                                        titulo, a_autor, n_autor, genero, año, isbn, fecha = datos
                                        hoja.cell(row=renglon, column=1).value=clave
                                        hoja.cell(row=renglon, column=2).value=titulo
                                        hoja.cell(row=renglon, column=3).value=a_autor
                                        hoja.cell(row=renglon, column=4).value=n_autor
                                        hoja.cell(row=renglon, column=5).value=genero
                                        hoja.cell(row=renglon, column=6).value=año
                                        hoja.cell(row=renglon, column=7).value=isbn
                                        hoja.cell(row=renglon, column=8).value=fecha
                                        renglon += 1
                                    nombre_xl.save('libros_filtro.xlsx')
                                    
                                elif exportar==3:
                                    print("Saliendo")
                                    break
                            break
                        else:
                            print('Opcion no valida')
                            continue
                elif menu==3:
                    break
                else:
                    print('Opcion no valida')
                    continue
        elif opcion==3:
            while True:
                ap_autor=input('Ingresa el apellido del autor:\n').upper()
                if ap_autor=='':
                    print('No se permite dejar vacios. Intente de nuevo')
                    continue
                else:
                    break
            while True:
                nom_autor=input('Ingresa el nombre del autor a registrar:\n').upper()
                if nom_autor=='':
                    print('No se permite dejar vacios. Intente de nuevo')
                    continue
                else:
                    break
            try:
                with sqlite3.connect(nombre_bd) as conn:
                    cursor = conn.cursor()
                    valores = (ap_autor, nom_autor)
                    cursor.execute("INSERT INTO autor (apellido, nombre) VALUES(?,?)", valores)
                    print(f"La clave asignada fue {cursor.lastrowid}")
            except Error as e:
                print (e)
            except Exception:
                print(f"Se produjo el siguiente error: {sys.exc_info()[0]}")
            finally:
                conn.close()
        
        elif opcion==4:
            while True:
                nom_genero=input('Ingresa el nombre del genero:\n').upper()
                if nom_genero=='':
                    print('No se permite dejar vacios. Intente de nuevo')
                    continue
                else:
                    break
            try:
                with sqlite3.connect(nombre_bd) as conn:
                    cursor = conn.cursor()
                    valores = (nom_genero,)
                    cursor.execute("INSERT INTO genero (nombre) VALUES(?)", valores)
                    print(f"La clave asignada fue {cursor.lastrowid}")
            except Error as e:
                print (e)
            except Exception:
                print(f"Se produjo el siguiente error: {sys.exc_info()[0]}")
            finally:
                conn.close()

        elif opcion==5:
            with open(nombre_csv,"w", newline="") as archivo:
                grabador = csv.writer(archivo)
                grabador.writerow(("Clave", "Titulo", "Apellido Autor", "Nombre Autor" "Genero", "Año", "ISBN", "Fecha de adquisicion"))
                grabador.writerows([(clave, datos[0], datos[1], datos[2], datos[3], datos[4], datos[5], datos[6]) for clave, datos in libros.items()])
            print(f"El archivo fue guardado con el nombre: {nombre_csv}")
            break
        else:
            print('Opcion no valida')